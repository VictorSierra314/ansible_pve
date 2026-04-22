#!/usr/bin/env python3

# Run like this
# python3 xslx_compliance_report_generation.py ansible_output.txt

import re
import sys
import os
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

# ---------------------------
# Configuration
# ---------------------------
SKIP_EXACT = {
    "Gathering Facts",
    "Ensure --limit is used",
    "Apply CIS Lv1 Server RHEL 9",
}

ROLE_PREFIX_RE = re.compile(r"^ansible-role-[^:]+:\s*")

TASK_RE = re.compile(r"^TASK \[(.*?)\]")
STATUS_RE = re.compile(r"^(ok|changed|failed|fatal): \[(.*?)\]")
UNREACHABLE_RE = re.compile(r"fatal: \[(.*?)\]: UNREACHABLE")

# ---------------------------
# Parse Ansible output
# ---------------------------
logfile = sys.argv[1]

# Extract base name without path or extension
input_basename = os.path.basename(logfile)
input_name, _ = os.path.splitext(input_basename)

tasks_order = []
current_task = None
hosts = set()
results = defaultdict(dict)

with open(logfile, "r", errors="ignore") as f:
    for line in f:
        line = line.strip()

        m_task = TASK_RE.match(line)
        if m_task:
            task = ROLE_PREFIX_RE.sub("", m_task.group(1))

            if task in SKIP_EXACT or "Gather" in task:
                current_task = None
                continue

            current_task = task
            if task not in tasks_order:
                tasks_order.append(task)
            continue

        if not current_task:
            continue

        m_status = STATUS_RE.match(line)
        if m_status:
            status, host = m_status.groups()
            hosts.add(host)

            if status == "changed":
                results[current_task][host] = "NOT COMPLIANT"
            elif status == "failed":
                results[current_task][host] = "FAILED"
            elif status == "ok":
                results[current_task].setdefault(host, "OK")
            continue

        m_unr = UNREACHABLE_RE.match(line)
        if m_unr:
            host = m_unr.group(1)
            hosts.add(host)
            results[current_task][host] = "UNREACHABLE"

hosts = sorted(hosts)

# ---------------------------
# Build Workbook
# ---------------------------
wb = Workbook()
ws = wb.active
ws.title = "Compliance Matrix"

# Header
ws.append(["Task Name"] + hosts)
for cell in ws[1]:
    cell.font = Font(bold=True)

# Data rows
for task in tasks_order:
    ws.append([task] + [results[task].get(h, "OK") for h in hosts])

last_task_row = ws.max_row

# ---------------------------
# Conditional formatting
# ---------------------------
OK_FILL = PatternFill("solid", start_color="C6EFCE")
NC_FILL = PatternFill("solid", start_color="FFC7CE")
WARN_FILL = PatternFill("solid", start_color="FFEB9C")

for r in range(2, last_task_row + 1):
    for c in range(2, ws.max_column + 1):
        cell = ws.cell(row=r, column=c)
        if cell.value == "OK":
            cell.fill = OK_FILL
        elif cell.value == "NOT COMPLIANT":
            cell.fill = NC_FILL
        elif cell.value in ("FAILED", "UNREACHABLE"):
            cell.fill = WARN_FILL

# Autosize
for col in range(1, ws.max_column + 1):
    ws.column_dimensions[get_column_letter(col)].width = 35

# ---------------------------
# Per-Host Summary Sheet
# ---------------------------
sum_ws = wb.create_sheet("Per-Host Compliance Summary")
sum_ws.append(["Host", "Checked Items", "Compliant (OK)", "Non-Compliant", "Compliance %"])
for cell in sum_ws[1]:
    cell.font = Font(bold=True)

for row_idx, host in enumerate(hosts, start=2):
    col_letter = get_column_letter(hosts.index(host) + 2)
    data_range = f"'Compliance Matrix'!{col_letter}$2:{col_letter}${last_task_row}"

    ok_formula = f'=COUNTIF({data_range},"OK")'
    nc_formula = f'=COUNTIF({data_range},"NOT COMPLIANT")'
    checked_formula = f"={ok_formula[1:]}+{nc_formula[1:]}"
    compliance_formula = f"=IF(B{row_idx}=0,0,C{row_idx}/B{row_idx})"

    sum_ws.append([
        host,
        checked_formula,
        ok_formula,
        nc_formula,
        compliance_formula,
    ])

for r in range(2, sum_ws.max_row + 1):
    sum_ws.cell(row=r, column=5).number_format = "0.00%"

for col in range(1, sum_ws.max_column + 1):
    sum_ws.column_dimensions[get_column_letter(col)].width = 35

# ---------------------------
# Save
# ---------------------------
today = datetime.today().strftime("%Y-%m-%d")
outfile = f"{today}_compliance_report_{input_name}.xlsx"
wb.save(outfile)

print(f"Compliance report written to {outfile}")